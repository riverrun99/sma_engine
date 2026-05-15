"""
local_writer.py — local file output, no cloud required.

Always-on fallback for subscribers who don't set up Google Sheets. Produces
two files in the configured output directory (default: ./output):

  signals_current.xlsx
    Overwritten every cycle. Contains a single sheet with the snapshot
    layout: top signal block, ranked leaderboard, system states block,
    regime cell.
    Mirrors what the Google Sheets "Current" tab shows.

  signals_log.csv
    Append-only. One row per scan cycle. New rows are added to the bottom;
    existing rows are never modified. Mirrors what the Google Sheets "Log"
    tab shows. CSV format chosen because it's universal — opens in Excel,
    Numbers, Google Sheets, any text editor.

If openpyxl is unavailable, the XLSX file is skipped and only the CSV is
written. The engine never crashes on output failures.
"""

from __future__ import annotations

import os
import csv
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Schema for the append-only log. Order matters — first row of the CSV.
LOG_COLUMNS = [
    "timestamp_utc",
    "top_ticker", "top_timeframe", "top_outfit_id", "top_outfit_periods",
    "top_entry_price", "top_offset", "top_hit_count", "top_convergence",
    "conv_ohlc", "conv_close", "conv_parm", "conv_timeseries",
    "sp500", "nasdaq", "dow", "vix", "svix", "russell2000", "russell3000", "semis",
    "regime_label",
]


class LocalWriter:
    """Writes Current snapshot (xlsx) + append-only Log (csv) to local disk."""

    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.log_path = self.output_dir / "signals_log.csv"
        self.current_xlsx_path = self.output_dir / "signals_current.xlsx"
        self._ensure_log_header()

    def _ensure_log_header(self) -> None:
        """Create the log CSV with header row if it doesn't exist."""
        if not self.log_path.exists():
            try:
                with open(self.log_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(LOG_COLUMNS)
            except Exception as e:
                logging.warning(f"LocalWriter: could not create log file: {e}")

    def _row_from_signal(
        self, signal: Optional[dict], systems: list, regime_label: Optional[str],
        ts: datetime,
    ) -> list:
        """Flatten signal + systems + regime into one row matching LOG_COLUMNS."""
        # Build a {system_name: state} lookup
        sys_state = {s.name: s.state for s in systems} if systems else {}

        if signal:
            conv = signal.get("convergence", {})
            periods_str = "/".join(str(p) for p in signal.get("outfit_periods", []))
            return [
                ts.isoformat(),
                signal.get("ticker", ""),
                signal.get("timeframe", ""),
                signal.get("outfit_id", ""),
                periods_str,
                signal.get("entry_price", ""),
                signal.get("offset_applied", ""),
                signal.get("hit_count", ""),
                conv.get("score", ""),
                int(bool(conv.get("ohlc_detection", False))),
                int(bool(conv.get("candle_close", False))),
                int(bool(conv.get("parm_price", False))),
                int(bool(conv.get("time_series", False))),
                sys_state.get("S&P 500", ""),
                sys_state.get("NASDAQ", ""),
                sys_state.get("Dow Jones", ""),
                sys_state.get("VIX", ""),
                sys_state.get("SVIX", ""),
                sys_state.get("Russell 2000", ""),
                sys_state.get("Russell 3000", ""),
                sys_state.get("Semiconductors", ""),
                regime_label or "",
            ]
        else:
            # No signal this cycle — still log the systems and regime
            return [
                ts.isoformat(),
                "", "", "", "", "", "", "", "",
                "", "", "", "",
                sys_state.get("S&P 500", ""),
                sys_state.get("NASDAQ", ""),
                sys_state.get("Dow Jones", ""),
                sys_state.get("VIX", ""),
                sys_state.get("SVIX", ""),
                sys_state.get("Russell 2000", ""),
                sys_state.get("Russell 3000", ""),
                sys_state.get("Semiconductors", ""),
                regime_label or "",
            ]

    def append_log_row(
        self,
        signal: Optional[dict],
        systems: list,
        regime_label: Optional[str] = None,
        ts: Optional[datetime] = None,
    ) -> None:
        """Append one row to the log CSV."""
        ts = ts or datetime.now(timezone.utc)
        row = self._row_from_signal(signal, systems, regime_label, ts)
        try:
            with open(self.log_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(row)
        except Exception as e:
            logging.warning(f"LocalWriter: append failed: {e}")

    def write_current_xlsx(
        self,
        signal: Optional[dict],
        top_n: list,
        systems: list,
        regime_label: Optional[str] = None,
        ts: Optional[datetime] = None,
    ) -> None:
        """Overwrite the Current xlsx with a snapshot of latest state."""
        if not OPENPYXL_AVAILABLE:
            return
        ts = ts or datetime.now(timezone.utc)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Current"

            bold = Font(bold=True)
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937",
                                       fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            row = 1

            # Header
            ws.cell(row=row, column=1, value="ELEMENT 47 — SMA ENGINE LIVE").font = Font(bold=True, size=14)
            row += 1
            ws.cell(row=row, column=1, value="Last update:")
            ws.cell(row=row, column=2, value=ts.isoformat())
            row += 2

            # Top signal block
            ws.cell(row=row, column=1, value="TOP SIGNAL").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            if signal:
                periods_str = "/".join(str(p) for p in signal.get("outfit_periods", []))
                pairs = [
                    ("Ticker", signal.get("ticker", "")),
                    ("Timeframe", signal.get("timeframe", "")),
                    ("Outfit", f"{periods_str} ({signal.get('outfit_name', '')})"),
                    ("Entry price", signal.get("entry_price", "")),
                    ("Offset", signal.get("offset_applied", "")),
                    ("Hit count", signal.get("hit_count", "")),
                    ("Convergence", signal.get("convergence", {}).get("score", "")),
                    ("Risk", signal.get("risk", "")),
                ]
                for label, value in pairs:
                    ws.cell(row=row, column=1, value=label).font = bold
                    ws.cell(row=row, column=2, value=value)
                    row += 1
            else:
                ws.cell(row=row, column=1, value="(no signal detected this cycle)")
                row += 1
            row += 1

            # Ranked leaderboard block
            ws.cell(row=row, column=1, value=f"TOP {len(top_n)} RANKED").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            headers = ["Rank", "Ticker", "TF", "Outfit", "Hits", "Conv", "Score"]
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col_idx, value=h)
                c.font = bold
            row += 1
            for entry in top_n:
                periods_str = "/".join(str(p) for p in entry.get("outfit_periods", []))
                ws.cell(row=row, column=1, value=entry.get("rank"))
                ws.cell(row=row, column=2, value=entry.get("ticker"))
                ws.cell(row=row, column=3, value=entry.get("timeframe"))
                ws.cell(row=row, column=4, value=periods_str)
                ws.cell(row=row, column=5, value=entry.get("hit_count"))
                ws.cell(row=row, column=6, value=entry.get("convergence"))
                ws.cell(row=row, column=7, value=entry.get("rank_score"))
                row += 1
            row += 1

            # System states block
            ws.cell(row=row, column=1, value="SYSTEM STATES").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            for h in ["System", "State", "Note"]:
                pass
            ws.cell(row=row, column=1, value="System").font = bold
            ws.cell(row=row, column=2, value="State").font = bold
            ws.cell(row=row, column=3, value="Note").font = bold
            row += 1
            for s in systems:
                ws.cell(row=row, column=1, value=s.name)
                ws.cell(row=row, column=2, value=s.state.upper())
                ws.cell(row=row, column=3, value=s.note)
                row += 1
            row += 1

            # Regime
            ws.cell(row=row, column=1, value="REGIME").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=2, value=regime_label or "(not yet computed)").font = bold

            # Column widths
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 30

            wb.save(self.current_xlsx_path)
        except Exception as e:
            logging.warning(f"LocalWriter: xlsx write failed: {e}")

    def write_cycle(
        self,
        signal: Optional[dict],
        top_n: list,
        systems: list,
        regime_label: Optional[str] = None,
        ts: Optional[datetime] = None,
    ) -> None:
        """Convenience: write Current xlsx + append Log row in one call."""
        ts = ts or datetime.now(timezone.utc)
        self.write_current_xlsx(signal, top_n, systems, regime_label, ts)
        self.append_log_row(signal, systems, regime_label, ts)
